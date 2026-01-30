import openpyxl
from datetime import datetime
from sqlalchemy.orm import Session
from app.db.session import SessionLocal
from app.excise.models import ExciseOwner, ExciseVehicle, ExciseRegistration, ExcisePayment

class ExciseImportService:
    def import_from_excel(self, file_path: str, progress_callback=None):
        """
        Imports data from an Excel file into the normalized database schema.
        Expected Columns (approximate match):
        - Registration No
        - Chassis No
        - Engine No
        - Make
        - Model
        - Year
        - Color
        - Owner Name
        - Father Name
        - CNIC
        - Address
        - City
        - Tax Paid Upto
        - Amount
        - Payment Date
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Map headers
            headers = {}
            for cell in sheet[1]:
                if cell.value:
                    headers[str(cell.value).strip().lower()] = cell.column - 1
            
            # minimal required columns
            required = ["chassis no", "owner name", "cnic"]
            missing = [req for req in required if req not in headers]
            # If "chassis no" is missing, maybe check for "chassis"
            if "chassis no" not in headers and "chassis" in headers:
                headers["chassis no"] = headers["chassis"]
            
            # Iterate rows
            db = SessionLocal()
            count = 0
            total_rows = sheet.max_row - 1
            
            try:
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                    data = self._row_to_dict(row, headers)
                    
                    if not data.get("chassis no") or not data.get("cnic"):
                        continue # Skip empty rows
                        
                    self._process_row(db, data)
                    count += 1
                    
                    if progress_callback and row_idx % 10 == 0:
                        progress_callback(row_idx, total_rows)
                
                db.commit()
                return True, f"Successfully imported {count} records."
                
            except Exception as e:
                db.rollback()
                return False, f"Error processing rows: {str(e)}"
            finally:
                db.close()
                
        except Exception as e:
            return False, f"File Error: {str(e)}"

    def _row_to_dict(self, row, headers):
        data = {}
        for key, col_idx in headers.items():
            if col_idx < len(row):
                data[key] = row[col_idx]
        return data

    def _process_row(self, db: Session, data: dict):
        # 1. Owner (Get or Create by CNIC)
        cnic = str(data.get("cnic", "")).strip()
        owner = db.query(ExciseOwner).filter(ExciseOwner.cnic == cnic).first()
        if not owner:
            owner = ExciseOwner(
                name=data.get("owner name"),
                father_name=data.get("father name"),
                cnic=cnic,
                address=data.get("address"),
                city=data.get("city")
            )
            db.add(owner)
            db.flush() # get ID
        
        # 2. Vehicle (Get or Create by Chassis)
        chassis = str(data.get("chassis no", "")).strip()
        vehicle = db.query(ExciseVehicle).filter(ExciseVehicle.chassis_number == chassis).first()
        if not vehicle:
            vehicle = ExciseVehicle(
                chassis_number=chassis,
                engine_number=data.get("engine no"),
                make=data.get("make"),
                model=data.get("model"),
                year=self._parse_int(data.get("year")),
                color=data.get("color"),
                horsepower=data.get("horsepower"),
                seating_capacity=self._parse_int(data.get("seating capacity"))
            )
            db.add(vehicle)
            db.flush()
            
        # 3. Registration (Get or Create by Reg No or Vehicle/Owner combo)
        reg_no = str(data.get("registration no") or f"TEMP-{chassis}").strip()
        registration = db.query(ExciseRegistration).filter(ExciseRegistration.registration_number == reg_no).first()
        
        if not registration:
            registration = ExciseRegistration(
                registration_number=reg_no,
                registration_date=self._parse_date(data.get("registration date")),
                token_tax_paid_upto=self._parse_date(data.get("tax paid upto")),
                owner_id=owner.id,
                vehicle_id=vehicle.id
            )
            db.add(registration)
            db.flush()
        else:
            # Update info if needed
            if data.get("tax paid upto"):
                registration.token_tax_paid_upto = self._parse_date(data.get("tax paid upto"))

        # 4. Payment (Add if amount exists)
        amount = self._parse_float(data.get("amount"))
        if amount and amount > 0:
            # Avoid duplicate payments? (Simple check: same date and amount for this reg)
            p_date = self._parse_date(data.get("payment date")) or datetime.now().date()
            
            exists = db.query(ExcisePayment).filter(
                ExcisePayment.registration_id == registration.id,
                ExcisePayment.amount == amount,
                ExcisePayment.payment_date == p_date
            ).first()
            
            if not exists:
                payment = ExcisePayment(
                    amount=amount,
                    payment_date=p_date,
                    challan_number=str(data.get("challan no") or ""),
                    payment_type=data.get("payment type", "Tax Payment"),
                    registration_id=registration.id
                )
                db.add(payment)

    def _parse_date(self, val):
        if not val:
            return None
        if isinstance(val, datetime):
            return val.date()
        try:
            return datetime.strptime(str(val), "%Y-%m-%d").date()
        except:
            return None

    def _parse_int(self, val):
        try:
            return int(float(str(val).strip()))
        except:
            return None
            
    def _parse_float(self, val):
        try:
            return float(str(val).strip())
        except:
            return 0.0

excise_service = ExciseImportService()
